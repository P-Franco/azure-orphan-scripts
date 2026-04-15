#!/usr/bin/env python3
"""
generate_excel_report.py

Runs the same Azure Resource Graph queries as orphan_report.py and exports
the results to a formatted Excel workbook with Summary, Production, and
Dev/QA/UAT sheets.

Usage:
  python3 generate_excel_report.py
  python3 generate_excel_report.py --output my_report.xlsx
  python3 generate_excel_report.py --subscription <id>
"""

import argparse
import json
import logging
import re
import sys
from datetime import datetime, timezone
from concurrent.futures import ThreadPoolExecutor, as_completed

from azure.identity import DefaultAzureCredential
from azure.mgmt.resourcegraph import ResourceGraphClient
from azure.mgmt.resourcegraph.models import QueryRequest, QueryRequestOptions
from azure.mgmt.subscription import SubscriptionClient
from azure.core.exceptions import HttpResponseError
from tenacity import retry, wait_exponential, stop_after_attempt, retry_if_exception_type
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger("excel-report")


# ── Environment classification ────────────────────────────────────────────
NON_PROD_KEYWORDS = [
    "dev", "development", "qa", "uat", "test", "staging", "sandbox",
    "lab", "pilot", "poc", "nonprod", "non-prod", "nonprd", "non-prd",
    "preprod", "pre-prod", "stg", "demo",
]


def classify_subscription(sub_name: str) -> str:
    name_lower = sub_name.lower()
    for kw in NON_PROD_KEYWORDS:
        if kw in name_lower:
            return "NON-PRODUCTION"
    return "PRODUCTION"


_ENV_TAG_KEYS = {"environment", "env"}
_NON_PROD_PATTERNS = [
    re.compile(r"(?<![a-z])" + re.escape(kw) + r"(?![a-z])")
    for kw in NON_PROD_KEYWORDS
]


def classify_resource(resource: dict, sub_envs: dict) -> str:
    """Classify a resource using tags, naming conventions, then subscription fallback."""
    tags = resource.get("tags") or {}
    if isinstance(tags, dict):
        for tk, tv in tags.items():
            if tk.lower() in _ENV_TAG_KEYS:
                val = str(tv).lower().strip()
                for kw in NON_PROD_KEYWORDS:
                    if kw in val:
                        return "NON-PRODUCTION"
                if any(p in val for p in ("prod", "prd")):
                    return "PRODUCTION"
    name_lower = resource.get("name", "").lower()
    rg_lower = resource.get("resourceGroup", "").lower()
    for pattern in _NON_PROD_PATTERNS:
        if pattern.search(name_lower) or pattern.search(rg_lower):
            return "NON-PRODUCTION"
    sub_id = resource.get("subscriptionId", "")
    return sub_envs.get(sub_id, "PRODUCTION")


# ── Resource Graph helper ────────────────────────────────────────────────────
@retry(
    retry=retry_if_exception_type(HttpResponseError),
    wait=wait_exponential(multiplier=1, min=2, max=30),
    stop=stop_after_attempt(3),
    reraise=True,
)
def run_query(graph_client, query, *, sub_ids=None, mgmt_group=None):
    results = []
    options = QueryRequestOptions(result_format="objectArray")
    kwargs = {"query": query, "options": options}
    if mgmt_group:
        kwargs["management_groups"] = [mgmt_group]
    elif sub_ids:
        kwargs["subscriptions"] = sub_ids
    request = QueryRequest(**kwargs)
    try:
        response = graph_client.resources(request)
        results.extend(response.data)
        while response.skip_token:
            options.skip_token = response.skip_token
            response = graph_client.resources(request)
            results.extend(response.data)
    except HttpResponseError as e:
        if e.status_code == 403:
            logger.warning(f"Insufficient permissions for query: {e.message}")
            return []
        logger.error(f"Resource Graph query failed: {e.message}")
        raise
    return results


# ── Queries (same as orphan_report.py) ────────────────────────────────────
QUERIES = {
    "App Service Plans with no apps": {
        "query": """Resources
| where type =~ 'microsoft.web/serverfarms'
| where properties.numberOfSites == 0
| project id=tolower(id), name, resourceGroup, location, subscriptionId, sku=tostring(sku.name), tags""",
        "cost": True, "section": "Compute",
    },
    "Availability Sets with no VMs": {
        "query": """Resources
| where type =~ 'microsoft.compute/availabilitysets'
| where properties.virtualMachines == '[]' or array_length(properties.virtualMachines) == 0
| project id=tolower(id), name, resourceGroup, location, subscriptionId, tags""",
        "cost": True, "section": "Compute",
    },
    "Unassociated Public IPs": {
        "query": """Resources
| where type =~ 'microsoft.network/publicipaddresses'
| where properties.ipConfiguration == '' or isnull(properties.ipConfiguration)
| where properties.natGateway == '' or isnull(properties.natGateway)
| project id=tolower(id), name, resourceGroup, location, subscriptionId, sku=tostring(sku.name), tags""",
        "cost": True, "section": "Networking",
    },
    "NICs not attached to a VM": {
        "query": """Resources
| where type =~ 'microsoft.network/networkinterfaces'
| where isnull(properties.virtualMachine) or properties.virtualMachine == ''
| where isnull(properties.privateEndpoint) or properties.privateEndpoint == ''
| project id=tolower(id), name, resourceGroup, location, subscriptionId, tags""",
        "cost": True, "section": "Networking",
    },
    "NSGs not associated with subnet or NIC": {
        "query": """Resources
| where type =~ 'microsoft.network/networksecuritygroups'
| where isnull(properties.networkInterfaces) or properties.networkInterfaces == '[]' or array_length(properties.networkInterfaces) == 0
| where isnull(properties.subnets) or properties.subnets == '[]' or array_length(properties.subnets) == 0
| project id=tolower(id), name, resourceGroup, location, subscriptionId, tags""",
        "cost": True, "section": "Networking",
    },
    "Load Balancers with empty backend pools": {
        "query": """Resources
| where type =~ 'microsoft.network/loadbalancers'
| where properties.backendAddressPools == '[]' or array_length(properties.backendAddressPools) == 0
| project id=tolower(id), name, resourceGroup, location, subscriptionId, sku=tostring(sku.name), tags""",
        "cost": True, "section": "Networking",
    },
    "Application Gateways with empty backend pools": {
        "query": """Resources
| where type =~ 'microsoft.network/applicationgateways'
| where properties.backendAddressPools == '[]' or array_length(properties.backendAddressPools) == 0
| project id=tolower(id), name, resourceGroup, location, subscriptionId, sku=tostring(sku.tier), tags""",
        "cost": True, "section": "Networking",
    },
    "VNet Gateways with no connections": {
        "query": """Resources
| where type =~ 'microsoft.network/virtualnetworkgateways'
| join kind=leftouter (
  Resources
  | where type =~ 'microsoft.network/connections'
  | mv-expand gw = pack_array(properties.virtualNetworkGateway1.id, properties.virtualNetworkGateway2.id)
  | project connectionGwId=tolower(tostring(gw))
) on $left.id == $right.connectionGwId
| where isnull(connectionGwId)
| project id=tolower(id), name, resourceGroup, location, subscriptionId, sku=tostring(properties.sku.name), tags""",
        "cost": True, "section": "Networking",
    },
    "Private Endpoints not connected to a resource": {
        "query": """Resources
| where type =~ 'microsoft.network/privateendpoints'
| where isnull(properties.privateLinkServiceConnections) or array_length(properties.privateLinkServiceConnections) == 0
| where isnull(properties.manualPrivateLinkServiceConnections) or array_length(properties.manualPrivateLinkServiceConnections) == 0
| project id=tolower(id), name, resourceGroup, location, subscriptionId, tags""",
        "cost": True, "section": "Networking",
    },
    "Route Tables not associated with a subnet": {
        "query": """Resources
| where type =~ 'microsoft.network/routetables'
| where isnull(properties.subnets) or properties.subnets == '[]' or array_length(properties.subnets) == 0
| project id=tolower(id), name, resourceGroup, location, subscriptionId, tags""",
        "cost": False, "section": "Networking",
    },
    "NAT Gateways not associated with a subnet": {
        "query": """Resources
| where type =~ 'microsoft.network/natgateways'
| where isnull(properties.subnets) or properties.subnets == '[]' or array_length(properties.subnets) == 0
| project id=tolower(id), name, resourceGroup, location, subscriptionId, sku=tostring(sku.name), tags""",
        "cost": True, "section": "Networking",
    },
    "Front Door WAF Policies not linked to a Front Door": {
        "query": """Resources
| where type =~ 'microsoft.network/frontdoorwebapplicationfirewallpolicies'
| where (isnull(properties.frontendEndpointLinks) or array_length(properties.frontendEndpointLinks) == 0)
| where (isnull(properties.securityPolicyLinks) or array_length(properties.securityPolicyLinks) == 0)
| project id=tolower(id), name, resourceGroup, location, subscriptionId, sku=tostring(sku.name), tags""",
        "cost": True, "section": "Networking",
    },
    "Traffic Manager Profiles with no endpoints": {
        "query": """Resources
| where type =~ 'microsoft.network/trafficmanagerprofiles'
| where properties.endpoints == '[]' or array_length(properties.endpoints) == 0
| project id=tolower(id), name, resourceGroup, location, subscriptionId, tags""",
        "cost": True, "section": "Networking",
    },
    "Virtual Networks with no subnets": {
        "query": """Resources
| where type =~ 'microsoft.network/virtualnetworks'
| where isnull(properties.subnets) or array_length(properties.subnets) == 0
| project id=tolower(id), name, resourceGroup, location, subscriptionId, tags""",
        "cost": False, "section": "Networking",
    },
    "Subnets without connected devices": {
        # Subnets have no independent resource ID in Cost Management; empty id
        # means cost lookup falls through to $0 (which is correct — subnets
        # don't incur standalone charges).
        "query": """Resources
| where type =~ 'microsoft.network/virtualnetworks'
| mv-expand subnet = properties.subnets
| where subnet.name !in~ ('GatewaySubnet', 'AzureFirewallSubnet', 'AzureFirewallManagementSubnet', 'AzureBastionSubnet', 'RouteServerSubnet')
| where (isnull(subnet.properties.ipConfigurations) or array_length(subnet.properties.ipConfigurations) == 0)
| where (isnull(subnet.properties.privateEndpoints) or array_length(subnet.properties.privateEndpoints) == 0)
| where (isnull(subnet.properties.delegations) or array_length(subnet.properties.delegations) == 0)
| extend subnetName = tostring(subnet.name)
| project id='', name=subnetName, resourceGroup, location, subscriptionId, sku=name, tags""",
        "cost": False, "section": "Networking",
    },
    "IP Groups not referenced by any firewall": {
        "query": """Resources
| where type =~ 'microsoft.network/ipgroups'
| where (isnull(properties.firewalls) or array_length(properties.firewalls) == 0)
| where (isnull(properties.firewallPolicies) or array_length(properties.firewallPolicies) == 0)
| project id=tolower(id), name, resourceGroup, location, subscriptionId, tags""",
        "cost": False, "section": "Networking",
    },
    "Private DNS Zones with no VNet links": {
        "query": """Resources
| where type =~ 'microsoft.network/privatednszones'
| where properties.numberOfVirtualNetworkLinks == 0
| project id=tolower(id), name, resourceGroup, location, subscriptionId, tags""",
        "cost": False, "section": "Networking",
    },
    "DDoS Protection Plans with no associated VNets": {
        "query": """Resources
| where type =~ 'microsoft.network/ddosprotectionplans'
| where isnull(properties.virtualNetworks) or properties.virtualNetworks == '[]' or array_length(properties.virtualNetworks) == 0
| project id=tolower(id), name, resourceGroup, location, subscriptionId, tags""",
        "cost": True, "section": "Networking",
    },
    "Unattached Managed Disks": {
        "query": """Resources
| where type =~ 'microsoft.compute/disks'
| where properties.diskState =~ 'Unattached'
| project id=tolower(id), name, resourceGroup, location, subscriptionId, sku=tostring(sku.name), tags""",
        "cost": True, "section": "Storage",
    },
    "SQL Elastic Pools with no databases": {
        "query": """Resources
| where type =~ 'microsoft.sql/servers/elasticpools'
| where isnull(properties.perDatabaseSettings) or properties.numberOfDatabases == 0
| project id=tolower(id), name, resourceGroup, location, subscriptionId, sku=tostring(sku.name), tags""",
        "cost": False, "section": "Database",
    },
    "Expired App Service Certificates": {
        "query": """Resources
| where type =~ 'microsoft.web/certificates'
| where properties.expirationDate < now()
| project id=tolower(id), name, resourceGroup, location, subscriptionId, sku=tostring(properties.expirationDate), tags""",
        "cost": False, "section": "Other",
    },
    "Disconnected API Connections": {
        "query": """Resources
| where type =~ 'microsoft.web/connections'
| where isnotnull(properties.statuses)
| where array_length(properties.statuses) > 0
| extend connStatus = tostring(properties.statuses[0]['status'])
| where connStatus !in~ ('Connected', 'Ready')
| project id=tolower(id), name, resourceGroup, location, subscriptionId, sku=connStatus, tags""",
        "cost": False, "section": "Other",
    },
}


# ── Estimated monthly cost per resource type (USD) ────────────────────────
COST_ESTIMATES = {
    "App Service Plans with no apps": 55.0,
    "Availability Sets with no VMs": 0.0,
    "Unassociated Public IPs": 3.60,
    "NICs not attached to a VM": 0.0,
    "NSGs not associated with subnet or NIC": 0.0,
    "Load Balancers with empty backend pools": 25.0,
    "Application Gateways with empty backend pools": 140.0,
    "VNet Gateways with no connections": 140.0,
    "Private Endpoints not connected to a resource": 7.30,
    "Route Tables not associated with a subnet": 0.0,
    "NAT Gateways not associated with a subnet": 32.0,
    "Front Door WAF Policies not linked to a Front Door": 5.0,
    "Traffic Manager Profiles with no endpoints": 0.36,
    "Virtual Networks with no subnets": 0.0,
    "Subnets without connected devices": 0.0,
    "IP Groups not referenced by any firewall": 0.0,
    "Private DNS Zones with no VNet links": 0.25,
    "DDoS Protection Plans with no associated VNets": 2944.0,
    "Unattached Managed Disks": 5.0,
    "SQL Elastic Pools with no databases": 150.0,
    "Expired App Service Certificates": 0.0,
    "Disconnected API Connections": 0.0,
    "Empty Resource Groups": 0.0,
}


# ── Empty Resource Groups ─────────────────────────────────────────────────
def find_empty_rgs(graph_client, **query_kwargs):
    all_rgs = run_query(
        graph_client,
        """ResourceContainers
| where type =~ 'microsoft.resources/subscriptions/resourcegroups'
| project name, resourceGroup=name, location, subscriptionId, tags""",
        **query_kwargs,
    )
    nonempty = run_query(
        graph_client,
        """Resources
| summarize count() by resourceGroup, subscriptionId
| project resourceGroup, subscriptionId""",
        **query_kwargs,
    )
    occupied = {
        (r["resourceGroup"].lower(), r["subscriptionId"].lower())
        for r in nonempty
    }
    return [
        rg for rg in all_rgs
        if (rg["resourceGroup"].lower(), rg["subscriptionId"].lower()) not in occupied
    ]


# ── Excel formatting constants ────────────────────────────────────────────
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
SECTION_FONT = Font(name="Calibri", bold=True, color="2F5496", size=11)
COST_YES = Font(name="Calibri", color="C00000")
COST_NO = Font(name="Calibri", color="548235")
PROD_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
NONPROD_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
THIN_BORDER = Border(
    bottom=Side(style="thin", color="D9D9D9"),
)
SUMMARY_HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
SUMMARY_HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=12)


def auto_width(ws):
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 3, 50)


def _format_tags(tags) -> str:
    """Convert tags dict to a readable string for Excel."""
    if not tags or not isinstance(tags, dict):
        return ""
    return json.dumps(tags, separators=(", ", ": "))


def _lookup_cost(r: dict, cat_name: str, cost_result):
    """Resolve (monthly_cost, rolling30d, last_billing_month, source) for a
    resource. Prefers real Cost Management data; falls back to the hardcoded
    per-category estimate if no billing history exists.

    Returns:
      (monthly_cost, rolling30d, last_billing_month, source)
      where `source` is one of: 'costManagement', 'estimate', 'estimate-zero-cm'
    """
    resource_id = str(r.get("id", "") or "").lower()
    estimate = float(COST_ESTIMATES.get(cat_name, 0.0))

    if cost_result is not None and resource_id:
        rec = cost_result.get_cost(resource_id)
        if rec is not None:
            # Prefer rolling 30d as canonical monthly figure.
            monthly = rec.rolling30d if rec.rolling30d > 0 else rec.last_billing_month
            if monthly > 0:
                return monthly, rec.rolling30d, rec.last_billing_month, "costManagement"
            # Real query matched but returned $0 — resource too new to bill.
            return estimate, rec.rolling30d, rec.last_billing_month, "estimate-zero-cm"

    return estimate, 0.0, 0.0, "estimate"


def write_detail_sheet(ws, title, rows_by_category, sub_names, sub_envs, env_filter, cost_result=None):
    """Write a detail sheet (Production or Dev/QA/UAT)."""
    # Headers
    headers = ["Category", "Section", "Resource Name", "Resource Group",
               "Location", "Subscription", "SKU / Detail", "Incurs Cost?",
               "Monthly Cost", "Rolling 30d (actual)", "Last Billing Month",
               "Cost Source", "Environment", "Tags"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="left")
    ws.freeze_panes = "A2"

    row_num = 2
    for cat_name, cfg, resources in rows_by_category:
        filtered = [r for r in resources
                    if classify_resource(r, sub_envs) == env_filter]
        if not filtered:
            continue

        for r in filtered:
            sub_id = r.get("subscriptionId", "")
            monthly, rolling, last_month, source = _lookup_cost(r, cat_name, cost_result)

            ws.cell(row=row_num, column=1, value=cat_name)
            ws.cell(row=row_num, column=2, value=cfg["section"])
            ws.cell(row=row_num, column=3, value=r.get("name", ""))
            ws.cell(row=row_num, column=4, value=r.get("resourceGroup", ""))
            ws.cell(row=row_num, column=5, value=r.get("location", ""))
            ws.cell(row=row_num, column=6, value=sub_names.get(sub_id, sub_id))
            ws.cell(row=row_num, column=7, value=r.get("sku", "") or "")

            cost_cell = ws.cell(row=row_num, column=8, value="Yes" if cfg["cost"] else "No")
            cost_cell.font = COST_YES if cfg["cost"] else COST_NO

            monthly_cell = ws.cell(row=row_num, column=9, value=round(monthly, 2))
            monthly_cell.number_format = '$#,##0.00'

            rolling_cell = ws.cell(row=row_num, column=10, value=round(rolling, 2))
            rolling_cell.number_format = '$#,##0.00'

            last_month_cell = ws.cell(row=row_num, column=11, value=round(last_month, 2))
            last_month_cell.number_format = '$#,##0.00'

            ws.cell(row=row_num, column=12, value=source)

            env = classify_resource(r, sub_envs)
            env_cell = ws.cell(row=row_num, column=13, value=env)
            env_cell.fill = PROD_FILL if env == "PRODUCTION" else NONPROD_FILL

            ws.cell(row=row_num, column=14, value=_format_tags(r.get("tags")))

            for c in range(1, 15):
                ws.cell(row=row_num, column=c).border = THIN_BORDER

            row_num += 1

    auto_width(ws)
    return row_num - 2  # data rows written


def main():
    parser = argparse.ArgumentParser(description="Generate Excel report of orphaned Azure resources")
    parser.add_argument("--subscription", "-s", help="Scope to a single subscription ID")
    parser.add_argument("--exclude-subscriptions", nargs="+", default=[],
                        help="Subscription IDs to exclude from scanning")
    parser.add_argument("--output", "-o", default="azure-orphan-report.xlsx",
                        help="Output Excel file (default: azure-orphan-report.xlsx)")
    parser.add_argument("--no-cost-data", action="store_true",
                        help="Skip Cost Management API enrichment and use hardcoded "
                             "estimates only. Use this if the caller lacks the "
                             "'Cost Management Reader' role.")
    parser.add_argument("--cost-cache-dir", default=".",
                        help="Directory for daily cost-cache-YYYYMMDD.json files (default: cwd)")
    parser.add_argument("--refresh-cost-data", action="store_true",
                        help="Force a fresh pull from Cost Management even if "
                             "a same-day cache exists.")
    args = parser.parse_args()

    print("Authenticating...")
    try:
        credential = DefaultAzureCredential()
        graph_client = ResourceGraphClient(credential)
        sub_client = SubscriptionClient(credential)
    except Exception as e:
        print(f"Authentication failed: {e}")
        return 1

    # ── Collect subscriptions ─────────────────────────────────────────────
    query_kwargs = {}
    if args.subscription:
        query_kwargs["sub_ids"] = [args.subscription]
        sub_names = {args.subscription: args.subscription}
    else:
        first_tenant = next(sub_client.tenants.list(), None)
        if first_tenant is None:
            print("No tenants found.")
            sys.exit(1)
        tenant_id = first_tenant.tenant_id
        query_kwargs["mgmt_group"] = tenant_id
        all_subs = run_query(
            graph_client,
            """ResourceContainers
| where type =~ 'microsoft.resources/subscriptions'
| where properties.state =~ 'Enabled'
| project subscriptionId, name""",
            mgmt_group=tenant_id,
        )
        excluded = set(args.exclude_subscriptions)
        sub_names = {s["subscriptionId"]: s["name"] for s in all_subs
                     if s["subscriptionId"] not in excluded}

    sub_envs = {sid: classify_subscription(sname) for sid, sname in sub_names.items()}
    print(f"Scope: {len(sub_names)} subscriptions")

    # ── Run queries ───────────────────────────────────────────────────────
    print("Running 24 resource queries in parallel...")
    category_results = {}

    def _run(cat_name, q):
        return cat_name, run_query(graph_client, q, **query_kwargs)

    with ThreadPoolExecutor(max_workers=8) as pool:
        futures = {
            pool.submit(_run, name, cfg["query"]): name
            for name, cfg in QUERIES.items()
        }
        empty_rg_future = pool.submit(find_empty_rgs, graph_client, **query_kwargs)

        for future in as_completed(futures):
            name = futures[future]
            try:
                _, data = future.result()
                category_results[name] = data
                if data:
                    print(f"  ✓ {name}: {len(data)} found")
            except Exception as e:
                print(f"  ✗ {name}: {e}")
                category_results[name] = []

        try:
            category_results["Empty Resource Groups"] = empty_rg_future.result()
            count = len(category_results["Empty Resource Groups"])
            if count:
                print(f"  ✓ Empty Resource Groups: {count} found")
        except Exception as e:
            print(f"  ✗ Empty Resource Groups: {e}")
            category_results["Empty Resource Groups"] = []

    # ── Cost Management enrichment (real cost data) ──────────────────────
    cost_result = None
    if not args.no_cost_data:
        try:
            from azure_cost_enrichment import enrich_costs

            subs_with_orphans = {
                r.get("subscriptionId", "")
                for rows in category_results.values()
                for r in rows
                if r.get("subscriptionId")
            }
            if subs_with_orphans:
                print("\nEnriching with Cost Management data...")
                cost_result = enrich_costs(
                    credential,
                    sorted(subs_with_orphans),
                    cache_dir=args.cost_cache_dir,
                    use_cache=not args.refresh_cost_data,
                )
                window = cost_result.rolling30d_window
                if window[0]:
                    print(f"  Rolling 30d window: {window[0]} → {window[1]}")
                print(
                    f"  {len(cost_result.subscriptions_queried)} subscription(s) "
                    f"queried, {len(cost_result.cost_map)} resource records, "
                    f"${cost_result.total_rolling30d:,.2f} total rolling 30d spend"
                )
                if cost_result.subscriptions_failed:
                    for sid, err in cost_result.subscriptions_failed.items():
                        name = sub_names.get(sid, sid)
                        print(f"  ⚠  {name}: {err}")
        except ImportError:
            print("  azure_cost_enrichment module not available — using estimates.")
            cost_result = None
        except Exception as e:
            print(f"  Cost enrichment failed ({e}) — falling back to estimates.")
            logger.warning(f"Cost enrichment failed: {e}")
            cost_result = None

    # ── Build ordered list of (category, config, rows) ────────────────────
    all_rows = []
    for cat_name, cfg in QUERIES.items():
        rows = category_results.get(cat_name, [])
        if rows:
            all_rows.append((cat_name, cfg, rows))

    empty_rgs = category_results.get("Empty Resource Groups", [])
    if empty_rgs:
        all_rows.append(("Empty Resource Groups", {"cost": False, "section": "Other"}, empty_rgs))

    # ── Counts ────────────────────────────────────────────────────────────
    total = sum(len(r) for _, _, r in all_rows)
    prod_total = sum(
        1 for _, _, rows in all_rows for r in rows
        if classify_resource(r, sub_envs) == "PRODUCTION"
    )
    nonprod_total = total - prod_total
    cost_total = sum(
        1 for _, cfg, rows in all_rows for r in rows
        if cfg["cost"]
    )

    print(f"\nTotal orphaned resources: {total}")
    print(f"  Production: {prod_total}")
    print(f"  Dev/QA/UAT: {nonprod_total}")
    print(f"  Cost-incurring: {cost_total}")

    # ── Generate Excel ────────────────────────────────────────────────────
    print(f"\nGenerating {args.output}...")
    wb = Workbook()

    # ── Summary sheet ─────────────────────────────────────────────────────
    ws_sum = wb.active
    ws_sum.title = "Summary"

    ws_sum.merge_cells("A1:D1")
    title_cell = ws_sum.cell(row=1, column=1, value="Azure Orphaned Resources Report")
    title_cell.font = Font(name="Calibri", bold=True, size=16, color="1F4E79")

    now = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")
    ws_sum.cell(row=2, column=1, value=f"Generated: {now}").font = Font(name="Calibri", size=10, italic=True)
    ws_sum.cell(row=3, column=1, value=f"Scope: {len(sub_names)} subscriptions (tenant-wide)").font = Font(name="Calibri", size=10, italic=True)

    # Cost Management footnote
    if cost_result is not None:
        window_from, window_to = cost_result.rolling30d_window
        ws_sum.cell(
            row=4, column=1,
            value=f"Cost data: Azure Cost Management AmortizedCost — "
                  f"rolling 30d window {window_from} → {window_to}"
        ).font = Font(name="Calibri", size=10, italic=True, color="548235")
    else:
        ws_sum.cell(
            row=4, column=1,
            value="Cost data: hardcoded per-category estimates "
                  "(Cost Management enrichment disabled or unavailable)"
        ).font = Font(name="Calibri", size=10, italic=True, color="C00000")

    # Summary table
    sum_headers = ["Category", "Section", "Production", "Dev/QA/UAT", "Total",
                   "Incurs Cost?", "Monthly Cost", "Rolling 30d (actual)",
                   "Last Billing Month"]
    for col, h in enumerate(sum_headers, 1):
        cell = ws_sum.cell(row=5, column=col, value=h)
        cell.font = SUMMARY_HEADER_FONT
        cell.fill = SUMMARY_HEADER_FILL

    row_num = 6
    grand_prod = 0
    grand_nonprod = 0
    grand_monthly = 0.0
    grand_rolling = 0.0
    grand_last_month = 0.0
    for cat_name, cfg, rows in all_rows:
        prod_count = sum(1 for r in rows if classify_resource(r, sub_envs) == "PRODUCTION")
        nonprod_count = len(rows) - prod_count
        grand_prod += prod_count
        grand_nonprod += nonprod_count

        # Sum per-resource monthly/rolling/last-month rather than
        # category_estimate * count — this picks up real costs where available.
        cat_monthly = 0.0
        cat_rolling = 0.0
        cat_last_month = 0.0
        for r in rows:
            monthly, rolling, last_month, _src = _lookup_cost(r, cat_name, cost_result)
            cat_monthly += monthly
            cat_rolling += rolling
            cat_last_month += last_month
        grand_monthly += cat_monthly
        grand_rolling += cat_rolling
        grand_last_month += cat_last_month

        ws_sum.cell(row=row_num, column=1, value=cat_name)
        ws_sum.cell(row=row_num, column=2, value=cfg["section"])
        ws_sum.cell(row=row_num, column=3, value=prod_count)
        ws_sum.cell(row=row_num, column=4, value=nonprod_count)
        ws_sum.cell(row=row_num, column=5, value=len(rows)).font = Font(bold=True)
        cost_cell = ws_sum.cell(row=row_num, column=6, value="Yes" if cfg["cost"] else "No")
        cost_cell.font = COST_YES if cfg["cost"] else COST_NO

        monthly_cell = ws_sum.cell(row=row_num, column=7, value=round(cat_monthly, 2))
        monthly_cell.number_format = '$#,##0.00'
        rolling_cell = ws_sum.cell(row=row_num, column=8, value=round(cat_rolling, 2))
        rolling_cell.number_format = '$#,##0.00'
        last_cell = ws_sum.cell(row=row_num, column=9, value=round(cat_last_month, 2))
        last_cell.number_format = '$#,##0.00'

        for c in range(1, 10):
            ws_sum.cell(row=row_num, column=c).border = THIN_BORDER
        row_num += 1

    # Totals row
    row_num += 1
    ws_sum.cell(row=row_num, column=1, value="TOTAL").font = Font(bold=True, size=12)
    ws_sum.cell(row=row_num, column=3, value=grand_prod).font = Font(bold=True, size=12)
    ws_sum.cell(row=row_num, column=4, value=grand_nonprod).font = Font(bold=True, size=12)
    ws_sum.cell(row=row_num, column=5, value=total).font = Font(bold=True, size=12)
    grand_monthly_cell = ws_sum.cell(row=row_num, column=7, value=round(grand_monthly, 2))
    grand_monthly_cell.font = Font(bold=True, size=12)
    grand_monthly_cell.number_format = '$#,##0.00'
    grand_rolling_cell = ws_sum.cell(row=row_num, column=8, value=round(grand_rolling, 2))
    grand_rolling_cell.font = Font(bold=True, size=12)
    grand_rolling_cell.number_format = '$#,##0.00'
    grand_last_cell = ws_sum.cell(row=row_num, column=9, value=round(grand_last_month, 2))
    grand_last_cell.font = Font(bold=True, size=12)
    grand_last_cell.number_format = '$#,##0.00'

    auto_width(ws_sum)

    # ── Production detail sheet ───────────────────────────────────────────
    ws_prod = wb.create_sheet("Production & SharedServices")
    write_detail_sheet(ws_prod, "Production", all_rows, sub_names, sub_envs, "PRODUCTION", cost_result)

    # ── Dev/QA/UAT detail sheet ───────────────────────────────────────────
    ws_dev = wb.create_sheet("Dev - QA - UAT")
    write_detail_sheet(ws_dev, "Dev/QA/UAT", all_rows, sub_names, sub_envs, "NON-PRODUCTION", cost_result)

    # ── All Resources sheet (flat, filterable) ────────────────────────────
    ws_all = wb.create_sheet("All Resources")
    headers = ["Category", "Section", "Resource Name", "Resource Group",
               "Location", "Subscription", "SKU / Detail", "Incurs Cost?",
               "Monthly Cost", "Rolling 30d (actual)", "Last Billing Month",
               "Cost Source", "Environment", "Tags"]
    for col, h in enumerate(headers, 1):
        cell = ws_all.cell(row=1, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
    ws_all.freeze_panes = "A2"
    ws_all.auto_filter.ref = "A1:N1"

    row_num = 2
    for cat_name, cfg, resources in all_rows:
        for r in resources:
            sub_id = r.get("subscriptionId", "")
            env = classify_resource(r, sub_envs)
            monthly, rolling, last_month, source = _lookup_cost(r, cat_name, cost_result)

            ws_all.cell(row=row_num, column=1, value=cat_name)
            ws_all.cell(row=row_num, column=2, value=cfg["section"])
            ws_all.cell(row=row_num, column=3, value=r.get("name", ""))
            ws_all.cell(row=row_num, column=4, value=r.get("resourceGroup", ""))
            ws_all.cell(row=row_num, column=5, value=r.get("location", ""))
            ws_all.cell(row=row_num, column=6, value=sub_names.get(sub_id, sub_id))
            ws_all.cell(row=row_num, column=7, value=r.get("sku", "") or "")

            cost_cell = ws_all.cell(row=row_num, column=8, value="Yes" if cfg["cost"] else "No")
            cost_cell.font = COST_YES if cfg["cost"] else COST_NO

            monthly_cell = ws_all.cell(row=row_num, column=9, value=round(monthly, 2))
            monthly_cell.number_format = '$#,##0.00'
            rolling_cell = ws_all.cell(row=row_num, column=10, value=round(rolling, 2))
            rolling_cell.number_format = '$#,##0.00'
            last_cell = ws_all.cell(row=row_num, column=11, value=round(last_month, 2))
            last_cell.number_format = '$#,##0.00'
            ws_all.cell(row=row_num, column=12, value=source)

            env_cell = ws_all.cell(row=row_num, column=13, value=env)
            env_cell.fill = PROD_FILL if env == "PRODUCTION" else NONPROD_FILL
            ws_all.cell(row=row_num, column=14, value=_format_tags(r.get("tags")))

            for c in range(1, 15):
                ws_all.cell(row=row_num, column=c).border = THIN_BORDER
            row_num += 1

    auto_width(ws_all)

    # ── Save ──────────────────────────────────────────────────────────────
    wb.save(args.output)
    print(f"\n✓ Report saved to: {args.output}")
    print(f"  Sheets: Summary | Production & SharedServices | Dev - QA - UAT | All Resources")
    if cost_result is not None:
        total_rows = sum(len(rows) for _, _, rows in all_rows)
        real_count = 0
        for cat_name, _, rows in all_rows:
            for r in rows:
                _, _, _, src = _lookup_cost(r, cat_name, cost_result)
                if src == "costManagement":
                    real_count += 1
        print(f"  {real_count} of {total_rows} rows backed by real Cost Management data")
    return 0


if __name__ == "__main__":
    sys.exit(main() or 0)
