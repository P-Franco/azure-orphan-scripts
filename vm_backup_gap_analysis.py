#!/usr/bin/env python3
"""
VM Backup Gap Analysis - single-run, read-only, parallel.

Reconciles live VM inventory (Azure Resource Graph) against backup protection
state across every Recovery Services vault in ONE tenant, and writes a set of
CSVs + a summary.md (+ an optional formatted .xlsx) ready for a client workbook.

WHY THIS EXISTS / WHY PYTHON:
  Resource Graph CANNOT see Recovery Services backup items, so the backup side
  must come from per-vault `az backup` calls. A sequential bash sweep paid the
  az-CLI cold-start (~1.5s) on every one of ~85 calls => ~2.5 min. This version
  discovers VMs *and* vaults via two Resource Graph queries, then fires the
  per-vault backup calls concurrently with a thread pool (explicit
  --subscription, never `az account set`) => ~15s.

BUCKETS (mutually exclusive + exhaustive over inventory):
  1. Protected     - VM exists, has a healthy protected item
  2. Missing       - VM exists, no protected item in any vault
  3. Job issues    - VM has a protected item but it is NOT healthy
  4. Orphan items  - protected item with no matching live VM (still billing)
  protected + missing + job-issues == total VMs.  orphans are separate.

  PLUS two non-bucket findings layered on top (do NOT change the counts above):
  - Failure history: last successful backup + days since (how long a VM has gone
    without a fresh recovery point) - read straight from each item's payload.
  - Protected-but-stale: a "healthy" VM whose newest recovery point is too old.

STRICTLY READ-ONLY. Only `az ... list/show/query`. Safe to run in production.
Deps: Azure CLI (logged in) + python3 stdlib. openpyxl is optional (Excel only).

Usage:
  az login
  python3 vm_backup_gap_analysis.py                       # the tenant you're in
  python3 vm_backup_gap_analysis.py --tenant <tenant-id>
  python3 vm_backup_gap_analysis.py --subscription <name|id>
  python3 vm_backup_gap_analysis.py -o /tmp/reports --workers 16 --stale-days 14
"""

import argparse
import concurrent.futures as cf
import csv
import datetime
import json
import os
import subprocess
import sys

# ---------------------------------------------------------------------------
# Tunables / constants
# ---------------------------------------------------------------------------

STALE_DAYS = 14               # newest recovery point older than this => "stale"
HEALTHY_HS = {"Passed", "ActionSuggested"}
HEALTHY_LBS = {"Completed", "CompletedWithWarnings"}

WARNINGS = []                 # collected az failures (strings); never fatal

# -- Branding (AHEAD client deliverable) ---------------------------------------
# AHEAD's brand is blue on white. BRAND_DARK / BRAND_ACCENT are professional
# blues - drop in the exact brand hex (or a logo) to finalize.
# 8-char ARGB (FF = opaque) so Excel never renders these washed out.
BRAND_FONT = "Calibri"
BRAND_DARK = "FF1F4E79"       # AHEAD deep blue - header bands + wordmark
BRAND_ACCENT = "FF4A90D9"     # AHEAD bright blue accent (approximate - confirm exact hex)
BRAND_LIGHT = "FFF4F4F4"      # zebra stripe
STATUS_BAD = "FFC00000"
STATUS_WARN = "FFBF8F00"
STATUS_GOOD = "FF375623"


# ---------------------------------------------------------------------------
# Small helpers
# ---------------------------------------------------------------------------

def log(*a):
    print(*a, file=sys.stderr, flush=True)


def section(title):
    log("\n" + "-" * 70)
    log("  " + title)
    log("-" * 70)


def az(args, timeout=180):
    """Run a read-only az command. Return (parsed_json_or_None, error_or_None).
    Never raises. On any failure, records a warning and returns (None, err)."""
    try:
        p = subprocess.run(["az", *args], capture_output=True, text=True, timeout=timeout)
    except Exception as e:  # noqa: BLE001 - resilience by design
        err = f"az {' '.join(args)} -> {e}"
        WARNINGS.append(err)
        return None, err
    if p.returncode != 0:
        err = f"az {' '.join(args)} -> {(p.stderr or '').strip()[:200]}"
        WARNINGS.append(err)
        return None, err
    try:
        return json.loads(p.stdout or "null"), None
    except json.JSONDecodeError as e:
        err = f"az {' '.join(args)} -> bad json: {e}"
        WARNINGS.append(err)
        return None, err


def parse_dt(iso):
    if not iso:
        return None
    try:
        return datetime.datetime.fromisoformat(str(iso).replace("Z", "+00:00"))
    except ValueError:
        return None


def days_since(iso):
    """Whole days between an ISO timestamp and now (UTC). None if unparseable."""
    dt = parse_dt(iso)
    if dt is None:
        return None
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=datetime.timezone.utc)
    return (datetime.datetime.now(datetime.timezone.utc) - dt).days


def storage_model(red_raw):
    """Extract storage redundancy from `az backup vault backup-properties show`.
    That command returns a LIST of config objects (vaultstorageconfig +
    vaultconfig) with the value nested under .properties - not a flat object."""
    def smt(o):
        if not isinstance(o, dict):
            return None
        p = o.get("properties") or {}
        return (o.get("storageModelType") or o.get("storageType")
                or p.get("storageModelType") or p.get("storageType"))
    if isinstance(red_raw, list):
        for o in red_raw:
            v = smt(o)
            if v:
                return v
        return "Unknown"
    return smt(red_raw) or "Unknown"


def dlabel(d):
    """Render a days-since value for output: a number, or 'never'."""
    return "never" if d is None else d


# ---------------------------------------------------------------------------
# Azure Resource Graph (inventory + vault discovery)
# ---------------------------------------------------------------------------

VM_KQL = """
Resources
| where type =~ 'microsoft.compute/virtualMachines'
| extend powerState = tostring(properties.extended.instanceView.powerState.displayStatus)
| extend osType = tostring(properties.storageProfile.osDisk.osType)
| extend vmSize = tostring(properties.hardwareProfile.vmSize)
| project name, resourceGroup, subscriptionId, location, osType, powerState, vmSize, id
| order by name asc
""".strip()

VAULT_KQL = """
resources
| where type =~ 'microsoft.recoveryservices/vaults'
| project name, resourceGroup, subscriptionId, location
| order by name asc
""".strip()


def arg_query(kql, sub_ids):
    """Run a Resource Graph query, following skip-token pagination."""
    rows, skip = [], None
    while True:
        args = ["graph", "query", "-q", kql, "--first", "1000",
                "--subscriptions", *sub_ids, "-o", "json"]
        if skip:
            args += ["--skip-token", skip]
        page, _ = az(args)
        if not page:
            break
        rows.extend(page.get("data", []) or [])
        skip = page.get("skipToken") or page.get("skip_token")
        if not skip:
            break
    return rows


# ---------------------------------------------------------------------------
# Per-vault backup state (parallel)
# ---------------------------------------------------------------------------

def fetch_vault(v):
    """Pull protected VM items + storage redundancy for one vault."""
    sub = v.get("subscriptionId", "")
    rg = v.get("resourceGroup", "")
    name = v.get("name", "")
    items, _ = az(["backup", "item", "list", "--subscription", sub, "-g", rg,
                   "--vault-name", name, "--backup-management-type", "AzureIaasVM",
                   "-o", "json"])
    red_raw, _ = az(["backup", "vault", "backup-properties", "show",
                     "--subscription", sub, "-n", name, "-g", rg, "-o", "json"])
    return {
        "vaultName": name, "resourceGroup": rg, "subscriptionId": sub,
        "location": v.get("location", ""), "redundancy": storage_model(red_raw),
        "items": items or [],
    }


# ---------------------------------------------------------------------------
# Reconciliation
# ---------------------------------------------------------------------------

def is_healthy(it):
    return (it["protectionState"] == "Protected"
            and it["healthStatus"] in HEALTHY_HS
            and it["lastBackupStatus"] in HEALTHY_LBS)


def issue_reason(it):
    if it["protectionState"] != "Protected":
        return "protectionState=" + (it["protectionState"] or "Unknown")
    if it["healthStatus"] not in HEALTHY_HS:
        return "healthStatus=" + (it["healthStatus"] or "Unknown")
    if it["lastBackupStatus"] not in HEALTHY_LBS:
        return "lastBackupStatus=" + (it["lastBackupStatus"] or "None")
    return "unknown"


def primary_sibling(name):
    low = name.lower()
    idx = low.rfind("02")
    if idx == -1:
        return None
    return name[:idx] + "01" + name[idx + 2:]


def redundancy_finding(r):
    rl = (r or "").lower()
    if "geo" in rl:
        return "OK - geo-redundant"
    if rl == "locallyredundant":
        return "FINDING: LRS - single datacenter, no regional/geo DR"
    if rl == "zoneredundant":
        return "REVIEW: ZRS - multi-AZ but single region (no geo DR)"
    return "Unknown - could not read backup-properties"


def reconcile(vms, vault_results, sub_names):
    for i, vm in enumerate(vms):
        vm["_idx"] = i
        vm["_key"] = (vm.get("name") or "").strip().lower()
        vm["subscriptionName"] = sub_names.get((vm.get("subscriptionId") or "").lower(), "")
    vms_by_id = {(vm.get("id") or "").strip().lower(): vm for vm in vms if vm.get("id")}
    vms_by_name = {}
    for vm in vms:
        vms_by_name.setdefault(vm["_key"], []).append(vm)

    # flatten protected items, annotated with their vault + failure history
    items = []
    for vr in vault_results:
        for it in vr["items"]:
            props = it.get("properties", it) or {}
            fn = props.get("friendlyName") or it.get("name") or ""
            items.append({
                "friendlyName": fn,
                "protectionState": props.get("protectionState") or "",
                "healthStatus": props.get("healthStatus") or "",
                "lastBackupStatus": props.get("lastBackupStatus") or "",
                "lastBackupTime": props.get("lastBackupTime") or "",
                "virtualMachineId": props.get("virtualMachineId") or "",
                "policyName": props.get("policyName") or "",
                "vaultName": vr["vaultName"],
                "vaultResourceGroup": vr["resourceGroup"],
                "vaultSubscriptionId": vr["subscriptionId"],
                "vaultRedundancy": vr["redundancy"],
                "lastSuccessfulBackup": props.get("lastRecoveryPoint") or "",
                "daysSinceLastSuccess": days_since(props.get("lastRecoveryPoint")),
            })

    # match items -> VMs (ARM id first, then case-insensitive name)
    items_for_vm = {}
    orphans = []
    for it in items:
        vmid = (it["virtualMachineId"] or "").strip().lower()
        fn = (it["friendlyName"] or "").strip().lower()
        targets = []
        if vmid and vmid in vms_by_id:
            targets = [vms_by_id[vmid]]
        elif fn and fn in vms_by_name:
            targets = vms_by_name[fn]
        if targets:
            for vm in targets:
                items_for_vm.setdefault(vm["_idx"], []).append(it)
        else:
            orphans.append(it)

    # bucket every VM
    protected, missing, job_issues = [], [], []
    for vm in vms:
        its = items_for_vm.get(vm["_idx"], [])
        if not its:
            missing.append(vm)
        elif any(is_healthy(x) for x in its):
            vm["_item"] = next(x for x in its if is_healthy(x))
            protected.append(vm)
        else:
            vm["_item"] = its[0]
            job_issues.append(vm)

    # pair-gap: unprotected *02 whose *01 sibling IS protected
    protected_names = {vm["_key"]: (vm.get("name") or "") for vm in protected}
    for vm in missing + job_issues:
        vm["_pairGap"], vm["_pairPrimary"] = "", ""
        prim = primary_sibling(vm.get("name") or "")
        if prim and prim.lower() in protected_names:
            vm["_pairGap"] = "Yes"
            vm["_pairPrimary"] = protected_names[prim.lower()]

    # protected-but-stale: healthy on paper, newest recovery point too old
    stale = []
    for vm in protected:
        d = vm.get("_item", {}).get("daysSinceLastSuccess")
        if d is None or d > STALE_DAYS:
            vm["_staleDays"] = d
            stale.append(vm)

    return dict(items=items, protected=protected, missing=missing,
                job_issues=job_issues, orphans=orphans, stale=stale,
                vms_by_id=vms_by_id, vms_by_name=vms_by_name)


# ---------------------------------------------------------------------------
# CSV output
# ---------------------------------------------------------------------------

def write_csv(out, name, header, rows):
    with open(os.path.join(out, name), "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(header)
        w.writerows(rows)


def build_tables(rec):
    """Return {csv_name: (header, rows)} - the single source for CSV + Excel."""
    protected, missing = rec["protected"], rec["missing"]
    job_issues, orphans, stale = rec["job_issues"], rec["orphans"], rec["stale"]
    items = rec["items"]
    pset = set(id(v) for v in protected)
    jset = set(id(v) for v in job_issues)

    def bucket(vm):
        return "Protected" if id(vm) in pset else ("JobIssue" if id(vm) in jset else "Missing")

    def matched(it):
        vmid = (it["virtualMachineId"] or "").strip().lower()
        fn = (it["friendlyName"] or "").strip().lower()
        return "Yes" if ((vmid and vmid in rec["vms_by_id"])
                         or (fn and fn in rec["vms_by_name"])) else "No"

    t = {}
    t["inventory.csv"] = (
        ["vmName", "powerState", "osType", "resourceGroup", "subscriptionId",
         "subscriptionName", "location", "vmSize", "backupBucket", "id"],
        [[vm.get("name", ""), vm.get("powerState", ""), vm.get("osType", ""),
          vm.get("resourceGroup", ""), vm.get("subscriptionId", ""),
          vm.get("subscriptionName", ""), vm.get("location", ""),
          vm.get("vmSize", ""), bucket(vm), vm.get("id", "")]
         for vm in _all_vms(rec)])

    t["missing-backup.csv"] = (
        ["vmName", "powerState", "osType", "resourceGroup", "subscriptionId",
         "subscriptionName", "location", "vmSize", "pairGap", "pairPrimary"],
        [[vm.get("name", ""), vm.get("powerState", ""), vm.get("osType", ""),
          vm.get("resourceGroup", ""), vm.get("subscriptionId", ""),
          vm.get("subscriptionName", ""), vm.get("location", ""), vm.get("vmSize", ""),
          vm.get("_pairGap", ""), vm.get("_pairPrimary", "")] for vm in missing])

    def jrow(vm):
        it = vm.get("_item", {})
        return [vm.get("name", ""), issue_reason(it), it.get("protectionState", ""),
                it.get("healthStatus", ""), it.get("lastBackupStatus", ""),
                it.get("lastBackupTime", ""), it.get("lastSuccessfulBackup", ""),
                dlabel(it.get("daysSinceLastSuccess")),
                it.get("vaultName", ""), it.get("vaultRedundancy", ""),
                vm.get("resourceGroup", ""), vm.get("subscriptionId", ""),
                vm.get("subscriptionName", ""), vm.get("location", ""),
                vm.get("vmSize", ""), vm.get("_pairGap", ""), vm.get("_pairPrimary", "")]

    t["job-health.csv"] = (
        ["vmName", "issueReason", "protectionState", "healthStatus", "lastBackupStatus",
         "lastBackupTime", "lastSuccessfulBackup", "daysSinceLastSuccess",
         "vaultName", "vaultRedundancy", "resourceGroup",
         "subscriptionId", "subscriptionName", "location", "vmSize", "pairGap", "pairPrimary"],
        [jrow(vm) for vm in job_issues])

    t["stale-protected.csv"] = (
        ["vmName", "lastSuccessfulBackup", "daysSinceLastSuccess", "healthStatus",
         "lastBackupStatus", "vaultName", "resourceGroup", "subscriptionName", "vmSize"],
        [[vm.get("name", ""), vm["_item"].get("lastSuccessfulBackup", ""),
          dlabel(vm.get("_staleDays")), vm["_item"].get("healthStatus", ""),
          vm["_item"].get("lastBackupStatus", ""), vm["_item"].get("vaultName", ""),
          vm.get("resourceGroup", ""), vm.get("subscriptionName", ""), vm.get("vmSize", "")]
         for vm in sorted(stale, key=lambda v: (v.get("_staleDays") is not None, v.get("_staleDays") or 0), reverse=True)])

    t["orphans.csv"] = (
        ["friendlyName", "protectionState", "healthStatus", "lastBackupStatus",
         "lastSuccessfulBackup", "daysSinceLastSuccess", "staleVirtualMachineId",
         "vaultName", "vaultResourceGroup", "vaultSubscriptionId", "vaultRedundancy"],
        [[it["friendlyName"], it["protectionState"], it["healthStatus"], it["lastBackupStatus"],
          it["lastSuccessfulBackup"], dlabel(it["daysSinceLastSuccess"]),
          it["virtualMachineId"], it["vaultName"], it["vaultResourceGroup"],
          it["vaultSubscriptionId"], it["vaultRedundancy"]] for it in orphans])

    t["protected-items.csv"] = (
        ["friendlyName", "matchedToLiveVM", "protectionState", "healthStatus",
         "lastBackupStatus", "lastBackupTime", "lastSuccessfulBackup",
         "daysSinceLastSuccess", "policyName", "vaultName",
         "vaultResourceGroup", "vaultSubscriptionId", "vaultRedundancy", "virtualMachineId"],
        [[it["friendlyName"], matched(it), it["protectionState"], it["healthStatus"],
          it["lastBackupStatus"], it["lastBackupTime"], it["lastSuccessfulBackup"],
          dlabel(it["daysSinceLastSuccess"]), it["policyName"],
          it["vaultName"], it["vaultResourceGroup"], it["vaultSubscriptionId"],
          it["vaultRedundancy"], it["virtualMachineId"]] for it in items])

    return t


def _all_vms(rec):
    # preserve name-sorted inventory order regardless of bucket lists
    seen, out = set(), []
    for vm in sorted(rec["protected"] + rec["missing"] + rec["job_issues"],
                     key=lambda v: (v.get("name") or "").lower()):
        if vm["_idx"] not in seen:
            seen.add(vm["_idx"])
            out.append(vm)
    return out


def vaults_table(vault_results):
    return (
        ["vaultName", "resourceGroup", "subscriptionId", "location", "redundancy",
         "redundancyFinding", "protectedItemCount"],
        [[m["vaultName"], m["resourceGroup"], m["subscriptionId"], m["location"],
          m["redundancy"], redundancy_finding(m["redundancy"]), len(m["items"])]
         for m in vault_results])


# ---------------------------------------------------------------------------
# summary.md
# ---------------------------------------------------------------------------

def write_summary(out, ts, tenant, scope_desc, vms, vault_results, rec):
    protected, missing = rec["protected"], rec["missing"]
    job_issues, orphans, stale = rec["job_issues"], rec["orphans"], rec["stale"]
    total = len(vms)
    n_prot, n_miss, n_job, n_orph = len(protected), len(missing), len(job_issues), len(orphans)
    n_vault = len(vault_results)
    coverage = (100.0 * n_prot / total) if total else 0.0
    recon_ok = (n_prot + n_miss + n_job == total)

    def red(m):
        return (m["redundancy"] or "").lower()
    single_region = [m for m in vault_results if red(m) in ("locallyredundant", "zoneredundant")]
    unknown_red = [m for m in vault_results if red(m).startswith("unknown") or not red(m).strip()]
    pair_gaps = [vm for vm in (missing + job_issues) if vm.get("_pairGap") == "Yes"]

    L = ["# VM Backup Gap Analysis", "",
         f"- **Run:** {ts}",
         f"- **Tenant:** {tenant or 'unknown'}",
         f"- **Scope:** {scope_desc}",
         f"- **Recovery Services vaults inspected:** {n_vault}", "",
         "## Headline numbers", "",
         "| Metric | Count |", "|---|---:|",
         f"| Total VMs in scope | {total} |",
         f"| Protected (healthy) | {n_prot} |",
         f"| Missing backup assignment | {n_miss} |",
         f"| Job issues / silent failures | {n_job} |",
         f"| Orphan vault items (no live VM) | {n_orph} |",
         f"| Protected-but-stale (RP > {STALE_DAYS}d) | {len(stale)} |",
         f"| **Backup coverage** | **{coverage:.1f}%** |", "",
         "## Reconciliation check", "",
         f"`protected ({n_prot}) + missing ({n_miss}) + job-issues ({n_job}) = "
         f"{n_prot + n_miss + n_job}` vs total `{total}` -> "
         f"{'OK' if recon_ok else 'MISMATCH (investigate)'}", "",
         "Buckets are mutually exclusive and exhaustive over live inventory. "
         "Protected-but-stale is an overlay on the Protected bucket (not a separate "
         "count). Orphan items sit outside inventory.", ""]

    L += ["## Silent failures (look protected, producing no recovery points)", ""]
    if job_issues:
        L += ["| VM | Reason | Last successful backup | Days since | Vault |",
              "|---|---|---|---:|---|"]
        for vm in job_issues:
            it = vm.get("_item", {})
            L.append(f"| {vm.get('name','')} | {issue_reason(it)} "
                     f"| {(it.get('lastSuccessfulBackup','') or 'never')[:10]} "
                     f"| {dlabel(it.get('daysSinceLastSuccess'))} "
                     f"| {it.get('vaultName','')} |")
    else:
        L.append("None.")
    L.append("")

    L += [f"## Protected but stale (newest recovery point > {STALE_DAYS} days old)", ""]
    if stale:
        L.append(f"{len(stale)} of {n_prot} 'protected' VMs are healthy on paper but have no fresh "
                 "restore point - effectively a gap:")
        L += ["", "| VM | Last good backup | Days stale | Vault |", "|---|---|---:|---|"]
        for vm in sorted(stale, key=lambda v: (v.get("_staleDays") is not None, v.get("_staleDays") or 0), reverse=True)[:20]:
            it = vm.get("_item", {})
            L.append(f"| {vm.get('name','')} | {(it.get('lastSuccessfulBackup','') or 'never')[:10]} "
                     f"| {dlabel(vm.get('_staleDays'))} | {it.get('vaultName','')} |")
        if len(stale) > 20:
            L.append(f"| ... and {len(stale) - 20} more | | | |")
    else:
        L.append(f"None - every protected VM has a recovery point within {STALE_DAYS} days.")
    L.append("")

    L += ["## Vault storage redundancy", ""]
    if single_region:
        L.append(f"{len(single_region)} of {n_vault} vault(s) have no geo-redundant backup storage:")
        L += ["", "| Vault | Subscription | Redundancy | Finding |", "|---|---|---|---|"]
        for m in single_region:
            L.append(f"| {m['vaultName']} | {m['subscriptionId']} | {m['redundancy']} "
                     f"| {redundancy_finding(m['redundancy'])} |")
        L += ["", "> LRS keeps backups in a single datacenter. A regional outage loses both the VM "
              "and its backups. GRS/RA-GZRS is the safer default for production recovery vaults."]
    elif unknown_red:
        L.append(f"No LRS/ZRS vaults found among those with readable redundancy, but storage "
                 f"redundancy could NOT be read for {len(unknown_red)} of {n_vault} vault(s) - "
                 "NOT asserting full geo-redundancy. See vaults.csv.")
    else:
        L.append("All inspected vaults use geo-redundant storage. No single-region/LRS finding.")
    if single_region and unknown_red:
        L += ["", f"> Note: redundancy unreadable for {len(unknown_red)} of {n_vault} vault(s)."]

    L += ["", "## Pair gap (per-VM assignment instead of policy enforcement)", ""]
    if pair_gaps:
        L.append(f"{len(pair_gaps)} secondary node(s) are unprotected while their `*01` primary "
                 "IS protected - a strong signal of per-VM backup assignment rather than tag-based "
                 "Azure Policy enforcement:")
        L += ["", "| Unprotected secondary | Protected primary | Subscription |", "|---|---|---|"]
        for vm in pair_gaps:
            L.append(f"| {vm.get('name','')} | {vm.get('_pairPrimary','')} "
                     f"| {vm.get('subscriptionName','') or vm.get('subscriptionId','')} |")
    else:
        L.append("No `*02`-unprotected / `*01`-protected pairs detected.")

    L += ["", "## Access warnings", ""]
    if WARNINGS:
        L.append(f"{len(WARNINGS)} az call(s) failed (likely missing Reader/Backup Reader). "
                 "Affected items may be under-counted:")
        L.append("")
        for w in WARNINGS[:25]:
            L.append(f"- `{w}`")
        if len(WARNINGS) > 25:
            L.append(f"- ... and {len(WARNINGS) - 25} more")
    else:
        L.append("None - all az calls succeeded.")
    L.append("")

    with open(os.path.join(out, "summary.md"), "w", encoding="utf-8") as f:
        f.write("\n".join(L))

    return dict(total=total, protected=n_prot, missing=n_miss, jobIssues=n_job,
                orphans=n_orph, stale=len(stale), vaults=n_vault,
                coverage=round(coverage, 1), warnings=len(WARNINGS))


# ---------------------------------------------------------------------------
# Excel workbook (optional - needs openpyxl)
# ---------------------------------------------------------------------------

def write_excel(out, ts, tenant, scope_desc, vault_results, rec, headline, client=""):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        log("[!] openpyxl not installed - skipping .xlsx (CSVs still written). "
            "pip install openpyxl to enable.")
        return None

    WHITE, GREY = "FFFFFFFF", "FF595959"
    hdr_font = Font(name=BRAND_FONT, bold=True, color=WHITE, size=11)
    hdr_fill = PatternFill("solid", fgColor=BRAND_DARK)
    zebra = PatternFill("solid", fgColor=BRAND_LIGHT)
    accent = PatternFill("solid", fgColor=BRAND_ACCENT)
    bad = Font(name=BRAND_FONT, color=STATUS_BAD, bold=True)
    warn = Font(name=BRAND_FONT, color=STATUS_WARN, bold=True)
    good = Font(name=BRAND_FONT, color=STATUS_GOOD, bold=True)
    border = Border(bottom=Side(style="thin", color="FFD9D9D9"))
    BULLET = chr(8226)

    wb = Workbook()
    footer_text = "AHEAD Confidential" + (f"   -   Prepared for {client}" if client else "")
    date_str = f"{ts[0:4]}-{ts[4:6]}-{ts[6:8]}" if len(ts) >= 8 else ts

    def style_footer(ws):
        ws.oddFooter.left.text = footer_text
        ws.oddFooter.right.text = "Page &P of &N"

    def bucket_font(v):
        return {"Protected": good, "Missing": bad, "JobIssue": warn}.get(v)

    def days_font(v):
        if v == "never":
            return bad
        return bad if isinstance(v, int) and v > STALE_DAYS else None

    def finding_font(v):
        s = str(v)
        return {"FINDING": bad, "REVIEW": warn, "OK": good}.get(s.split(":")[0].split(" ")[0])

    def data_sheet(title, header, rows, color_rules=None):
        ws = wb.create_sheet(title[:31])
        for c, h in enumerate(header, 1):
            cell = ws.cell(row=1, column=c, value=h)
            cell.font, cell.fill = hdr_font, hdr_fill
            cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[1].height = 20
        for r, row in enumerate(rows, 2):
            for c, val in enumerate(row, 1):
                cell = ws.cell(row=r, column=c, value=val)
                cell.border = border
                if r % 2 == 0:
                    cell.fill = zebra
            for idx, fn in (color_rules or {}).items():
                f = fn(row[idx])
                if f:
                    ws.cell(row=r, column=idx + 1).font = f
        ws.freeze_panes = "A2"
        if rows:
            ws.auto_filter.ref = f"A1:{get_column_letter(len(header))}{len(rows) + 1}"
        for col in ws.columns:
            w = max((len(str(c.value)) for c in col if c.value is not None), default=8)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(w + 3, 60)
        style_footer(ws)
        return ws

    # ---- Cover / Summary ----
    sm = wb.active
    sm.title = "Summary"
    sm.sheet_view.showGridLines = False
    sm.column_dimensions["A"].width = 3
    sm.column_dimensions["B"].width = 46
    sm.column_dimensions["C"].width = 16
    sm.column_dimensions["D"].width = 34
    style_footer(sm)

    sm["B2"] = "AHEAD"
    sm["B2"].font = Font(name=BRAND_FONT, bold=True, size=30, color=BRAND_DARK)
    for col in ("B", "C", "D"):
        sm[f"{col}3"].fill = accent
    sm.row_dimensions[3].height = 6
    sm["B5"] = "Azure VM Backup Gap Analysis"
    sm["B5"].font = Font(name=BRAND_FONT, bold=True, size=16, color=BRAND_DARK)
    if client:
        sm["B6"] = f"Prepared for {client}"
        sm["B6"].font = Font(name=BRAND_FONT, size=11, italic=True, color=GREY)

    r = 8
    for label, val in [("Report date", date_str), ("Azure tenant", tenant),
                       ("Scope", scope_desc), ("Prepared by", "AHEAD")]:
        sm.cell(row=r, column=2, value=label).font = Font(name=BRAND_FONT, bold=True, size=10, color=GREY)
        sm.cell(row=r, column=3, value=val).font = Font(name=BRAND_FONT, size=10)
        r += 1

    r += 1
    sm.cell(row=r, column=2, value="Findings").font = Font(name=BRAND_FONT, bold=True, size=13, color=BRAND_DARK)
    r += 1
    cov = headline["coverage"]
    metrics = [
        ("Total VMs in scope", headline["total"], None),
        ("Protected (healthy)", headline["protected"], good),
        ("Missing backup assignment", headline["missing"], bad if headline["missing"] else good),
        ("Silent failures (look protected, aren't)", headline["jobIssues"], bad if headline["jobIssues"] else good),
        ("Orphan vault items (no live VM)", headline["orphans"], warn if headline["orphans"] else good),
        (f"Protected but stale (RP > {STALE_DAYS}d)", headline["stale"], bad if headline["stale"] else good),
        ("Backup coverage", f"{cov}%", good if cov >= 90 else (warn if cov >= 75 else bad)),
    ]
    for label, val, fnt in metrics:
        lc = sm.cell(row=r, column=2, value=label)
        lc.font, lc.border = Font(name=BRAND_FONT, size=11), border
        vc = sm.cell(row=r, column=3, value=val)
        vc.font = fnt or Font(name=BRAND_FONT, bold=True, size=11)
        vc.border = border
        r += 1

    # Top findings (auto-generated from the numbers)
    r += 1
    sm.cell(row=r, column=2, value="Top findings").font = Font(name=BRAND_FONT, bold=True, size=13, color=BRAND_DARK)
    r += 1
    lines = []
    if rec["job_issues"]:
        worst = max(rec["job_issues"], key=lambda v: (v["_item"].get("daysSinceLastSuccess") or 0))
        lines.append(f"{worst['name']} appears protected but has had no successful backup in "
                     f"{worst['_item'].get('daysSinceLastSuccess')} days.")
    lines.append(f"{headline['missing']} of {headline['total']} VMs have no backup assignment "
                 f"({cov}% coverage).")
    pg = [v for v in (rec["missing"] + rec["job_issues"]) if v.get("_pairGap") == "Yes"]
    if pg:
        lines.append(f"{len(pg)} secondary nodes are unprotected while their primary is backed up - "
                     "points to per-VM assignment rather than policy enforcement.")
    if headline["orphans"]:
        lines.append(f"{headline['orphans']} orphaned backup items remain for VMs that no longer exist.")
    for ln in lines:
        sm.cell(row=r, column=2, value=f"{BULLET}  {ln}").font = Font(name=BRAND_FONT, size=10, color="FF404040")
        r += 1

    # ---- Data sheets ----
    tables = build_tables(rec)
    data_sheet("Missing Backup", *tables["missing-backup.csv"])
    data_sheet("Silent Failures", *tables["job-health.csv"], color_rules={7: days_font})
    data_sheet("Stale Protected", *tables["stale-protected.csv"], color_rules={2: days_font})
    data_sheet("Pair Gaps", ["vmName", "pairGap", "pairPrimary", "subscriptionName"],
               [[vm.get("name", ""), "Yes", vm.get("_pairPrimary", ""), vm.get("subscriptionName", "")]
                for vm in (rec["missing"] + rec["job_issues"]) if vm.get("_pairGap") == "Yes"])
    data_sheet("Orphans", *tables["orphans.csv"])
    data_sheet("Vaults", *vaults_table(vault_results), color_rules={5: finding_font})
    data_sheet("Full Inventory", *tables["inventory.csv"], color_rules={8: bucket_font})

    path = os.path.join(out, "vm-backup-gap-analysis.xlsx")
    wb.save(path)
    return path


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    global STALE_DAYS
    ap = argparse.ArgumentParser(description="VM backup gap analysis (read-only).")
    ap.add_argument("-t", "--tenant", default="", help="tenant id (default: logged-in tenant)")
    ap.add_argument("-s", "--subscription", default="", help="limit to one subscription (name or id)")
    ap.add_argument("-o", "--output-dir", default=".", help="base output dir")
    ap.add_argument("--client", default="", help="client name for the report cover (e.g. 'Old Republic')")
    ap.add_argument("--workers", type=int, default=12, help="parallel vault fetches (default 12)")
    ap.add_argument("--stale-days", type=int, default=STALE_DAYS,
                    help=f"flag protected VMs whose newest recovery point is older than this (default {STALE_DAYS})")
    args = ap.parse_args()
    STALE_DAYS = args.stale_days

    section("Preflight")
    acct, _ = az(["account", "show", "-o", "json"])
    if not acct:
        log("[x] No active Azure session. Run 'az login' first.")
        sys.exit(1)
    login_tenant = acct.get("tenantId", "")
    log(f"[+] Authenticated as: {acct.get('user', {}).get('name', 'unknown')}")
    log(f"[+] Logged-in tenant: {login_tenant}")

    ext, _ = az(["extension", "show", "--name", "resource-graph", "-o", "json"])
    if not ext:
        log("[!] resource-graph extension missing - installing (read-only, local)...")
        WARNINGS.clear()
        subprocess.run(["az", "extension", "add", "--name", "resource-graph"],
                       capture_output=True, text=True)

    all_subs, _ = az(["account", "list", "-o", "json"])
    all_subs = all_subs or []
    sub_names = {(s.get("id") or "").lower(): s.get("name", "") for s in all_subs}

    if args.subscription:
        flt = args.subscription.strip().lower()
        match = next((s for s in all_subs
                      if s.get("id", "").lower() == flt or s.get("name", "").lower() == flt), None)
        if not match:
            log(f"[x] Subscription '{args.subscription}' not found.")
            sys.exit(1)
        sub_ids = [match["id"]]
        target_tenant = match.get("tenantId", login_tenant)
        scope_desc = f"subscription = {match.get('name', match['id'])}"
        log(f"[+] Scope: single subscription - {match.get('name')} ({match['id']})")
    else:
        target_tenant = args.tenant or login_tenant
        sub_ids = [s["id"] for s in all_subs
                   if s.get("state") == "Enabled" and s.get("tenantId") == target_tenant]
        if not sub_ids:
            log(f"[x] No enabled subscriptions in tenant {target_tenant}.")
            sys.exit(1)
        scope_desc = f"tenant {target_tenant} (all enabled subscriptions in this tenant)"
        log(f"[+] Target tenant: {target_tenant}")
        log(f"[+] Scope: {len(sub_ids)} enabled subscription(s)")

    section("VM inventory + vault discovery (Resource Graph)")
    vms = arg_query(VM_KQL, sub_ids)
    vaults = arg_query(VAULT_KQL, sub_ids)
    log(f"[+] {len(vms)} VM(s), {len(vaults)} Recovery Services vault(s)")

    section(f"Backup state - {len(vaults)} vault(s), {args.workers} workers")
    vault_results = []
    with cf.ThreadPoolExecutor(max_workers=max(1, args.workers)) as ex:
        for vr in ex.map(fetch_vault, vaults):
            vault_results.append(vr)
            log(f"[+]   {vr['vaultName']} [{vr['redundancy']}] - {len(vr['items'])} item(s)")

    section("Reconcile & report")
    rec = reconcile(vms, vault_results, sub_names)

    ts = datetime.datetime.now().strftime("%Y%m%d-%H%M%S")
    out = os.path.join(args.output_dir.rstrip("/"), f"vm-backup-gap-analysis-{ts}")
    raw = os.path.join(out, "raw")
    os.makedirs(raw, exist_ok=True)
    json.dump(vms, open(os.path.join(raw, "inventory.json"), "w"))
    json.dump(vault_results, open(os.path.join(raw, "vault_results.json"), "w"))
    if WARNINGS:
        open(os.path.join(raw, "warnings.log"), "w").write("\n".join(WARNINGS))

    tables = build_tables(rec)
    for name, (header, rows) in tables.items():
        write_csv(out, name, header, rows)
    write_csv(out, "vaults.csv", *vaults_table(vault_results))
    headline = write_summary(out, ts, target_tenant, scope_desc, vms, vault_results, rec)
    xlsx = write_excel(out, ts, target_tenant, scope_desc, vault_results, rec, headline, client=args.client)

    print(json.dumps(headline))
    section("Done")
    log(f"[+] Output folder: {out}")
    if xlsx:
        log(f"[+] Excel workbook: {os.path.basename(xlsx)}")
    if WARNINGS:
        log(f"[!] {len(WARNINGS)} az call(s) failed - see raw/warnings.log.")
    log("[+] Open summary.md (or the .xlsx) first.")


if __name__ == "__main__":
    main()
