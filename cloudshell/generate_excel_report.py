#!/usr/bin/env python3
"""
generate_excel_report.py

Runs the same Azure Resource Graph queries as orphan_report.py and exports
the results to a formatted Excel workbook with Summary, Production, and
Dev/QA/UAT sheets.

Usage:
  python3 generate_excel_report.py
  python3 generate_excel_report.py --output my_report.xlsx
  python3 generate_excel_report.py --subscription <id>
"""

import argparse
import json
import logging
import sys
from datetime import datetime, timezone
from concurrent.futures import ThreadPoolExecutor, as_completed

from azure.identity import DefaultAzureCredential
from azure.mgmt.resourcegraph import ResourceGraphClient
from azure.mgmt.subscription import SubscriptionClient
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Single source of truth for queries, estimates, classification, ARG access,
# and safety filters is orphan_report.py — this module only does Excel.
from orphan_report import (
    COST_ESTIMATES,
    QUERIES,
    apply_subscription_exclusions,
    classify_resource,
    classify_subscription,
    filter_do_not_delete,
    find_empty_rgs,
    run_query,
)

logger = logging.getLogger("excel-report")


# ── Excel formatting constants ────────────────────────────────────────────
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
SECTION_FONT = Font(name="Calibri", bold=True, color="2F5496", size=11)
COST_YES = Font(name="Calibri", color="C00000")
COST_NO = Font(name="Calibri", color="548235")
PROD_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
NONPROD_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
THIN_BORDER = Border(
    bottom=Side(style="thin", color="D9D9D9"),
)
SUMMARY_HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
SUMMARY_HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=12)


def auto_width(ws):
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 3, 50)


def _format_tags(tags) -> str:
    """Convert tags dict to a readable string for Excel."""
    if not tags or not isinstance(tags, dict):
        return ""
    return json.dumps(tags, separators=(", ", ": "))


def _lookup_cost(r: dict, cat_name: str, cost_result):
    """Resolve (monthly_cost, rolling30d, last_billing_month, source) for a
    resource. Prefers real Cost Management data; falls back to the hardcoded
    per-category estimate if no billing history exists.

    Returns:
      (monthly_cost, rolling30d, last_billing_month, source)
      where `source` is one of: 'costManagement', 'estimate', 'estimate-zero-cm'
    """
    resource_id = str(r.get("id", "") or "").lower()
    estimate = float(COST_ESTIMATES.get(cat_name, 0.0))

    if cost_result is not None and resource_id:
        rec = cost_result.get_cost(resource_id)
        if rec is not None:
            # Prefer rolling 30d as canonical monthly figure.
            monthly = rec.rolling30d if rec.rolling30d > 0 else rec.last_billing_month
            if monthly > 0:
                return monthly, rec.rolling30d, rec.last_billing_month, "costManagement"
            # Real query matched but returned $0 — resource too new to bill.
            return estimate, rec.rolling30d, rec.last_billing_month, "estimate-zero-cm"

    return estimate, 0.0, 0.0, "estimate"


def write_detail_sheet(ws, title, rows_by_category, sub_names, sub_envs, env_filter, cost_result=None):
    """Write a detail sheet (Production or Dev/QA/UAT)."""
    # Headers
    headers = ["Category", "Section", "Resource Name", "Resource Group",
               "Location", "Subscription", "SKU / Detail", "Incurs Cost?",
               "Monthly Cost", "Rolling 30d (actual)", "Last Billing Month",
               "Cost Source", "Environment", "Tags"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="left")
    ws.freeze_panes = "A2"

    row_num = 2
    for cat_name, cfg, resources in rows_by_category:
        filtered = [r for r in resources
                    if classify_resource(r, sub_envs) == env_filter]
        if not filtered:
            continue

        for r in filtered:
            sub_id = r.get("subscriptionId", "")
            monthly, rolling, last_month, source = _lookup_cost(r, cat_name, cost_result)

            ws.cell(row=row_num, column=1, value=cat_name)
            ws.cell(row=row_num, column=2, value=cfg["section"])
            ws.cell(row=row_num, column=3, value=r.get("name", ""))
            ws.cell(row=row_num, column=4, value=r.get("resourceGroup", ""))
            ws.cell(row=row_num, column=5, value=r.get("location", ""))
            ws.cell(row=row_num, column=6, value=sub_names.get(sub_id, sub_id))
            ws.cell(row=row_num, column=7, value=r.get("sku", "") or "")

            cost_cell = ws.cell(row=row_num, column=8, value="Yes" if cfg["cost"] else "No")
            cost_cell.font = COST_YES if cfg["cost"] else COST_NO

            monthly_cell = ws.cell(row=row_num, column=9, value=round(monthly, 2))
            monthly_cell.number_format = '$#,##0.00'

            rolling_cell = ws.cell(row=row_num, column=10, value=round(rolling, 2))
            rolling_cell.number_format = '$#,##0.00'

            last_month_cell = ws.cell(row=row_num, column=11, value=round(last_month, 2))
            last_month_cell.number_format = '$#,##0.00'

            ws.cell(row=row_num, column=12, value=source)

            env = classify_resource(r, sub_envs)
            env_cell = ws.cell(row=row_num, column=13, value=env)
            env_cell.fill = PROD_FILL if env == "PRODUCTION" else NONPROD_FILL

            ws.cell(row=row_num, column=14, value=_format_tags(r.get("tags")))

            for c in range(1, 15):
                ws.cell(row=row_num, column=c).border = THIN_BORDER

            row_num += 1

    auto_width(ws)
    return row_num - 2  # data rows written


def main():
    parser = argparse.ArgumentParser(description="Generate Excel report of orphaned Azure resources")
    parser.add_argument("--subscription", "-s", help="Scope to a single subscription ID")
    parser.add_argument("--exclude-subscriptions", nargs="+", default=[],
                        help="Subscription IDs to exclude from scanning")
    parser.add_argument("--output", "-o", default="azure-orphan-report.xlsx",
                        help="Output Excel file (default: azure-orphan-report.xlsx)")
    parser.add_argument("--no-cost-data", action="store_true",
                        help="Skip Cost Management API enrichment and use hardcoded "
                             "estimates only. Use this if the caller lacks the "
                             "'Cost Management Reader' role.")
    parser.add_argument("--cost-cache-dir", default=".",
                        help="Directory for daily cost-cache-YYYYMMDD.json files (default: cwd)")
    parser.add_argument("--refresh-cost-data", action="store_true",
                        help="Force a fresh pull from Cost Management even if "
                             "a same-day cache exists.")
    args = parser.parse_args()

    print("Authenticating...")
    try:
        credential = DefaultAzureCredential()
        graph_client = ResourceGraphClient(credential)
        sub_client = SubscriptionClient(credential)
    except Exception as e:
        print(f"Authentication failed: {e}")
        return 1

    # ── Collect subscriptions ─────────────────────────────────────────────
    query_kwargs = {}
    excluded = set(args.exclude_subscriptions)
    if args.subscription:
        query_kwargs["sub_ids"] = [args.subscription]
        sub_names = {args.subscription: args.subscription}
    else:
        first_tenant = next(sub_client.tenants.list(), None)
        if first_tenant is None:
            print("No tenants found.")
            sys.exit(1)
        tenant_id = first_tenant.tenant_id
        query_kwargs["mgmt_group"] = tenant_id
        all_subs = run_query(
            graph_client,
            """ResourceContainers
| where type =~ 'microsoft.resources/subscriptions'
| where properties.state =~ 'Enabled'
| project subscriptionId, name""",
            mgmt_group=tenant_id,
        )
        sub_names = {s["subscriptionId"]: s["name"] for s in all_subs
                     if s["subscriptionId"] not in excluded}

    sub_envs = {sid: classify_subscription(sname) for sid, sname in sub_names.items()}
    print(f"Scope: {len(sub_names)} subscriptions")

    # ── Run queries ───────────────────────────────────────────────────────
    print(f"Running {len(QUERIES) + 1} resource queries in parallel...")
    category_results = {}

    def _run(cat_name, q):
        return cat_name, run_query(graph_client, q, **query_kwargs)

    with ThreadPoolExecutor(max_workers=8) as pool:
        futures = {
            pool.submit(_run, name, cfg["query"]): name
            for name, cfg in QUERIES.items()
        }
        empty_rg_future = pool.submit(find_empty_rgs, graph_client, **query_kwargs)

        for future in as_completed(futures):
            name = futures[future]
            try:
                _, data = future.result()
                category_results[name] = data
                if data:
                    print(f"  ✓ {name}: {len(data)} found")
            except Exception as e:
                print(f"  ✗ {name}: {e}")
                category_results[name] = []

        try:
            category_results["Empty Resource Groups"] = empty_rg_future.result()
            count = len(category_results["Empty Resource Groups"])
            if count:
                print(f"  ✓ Empty Resource Groups: {count} found")
        except Exception as e:
            print(f"  ✗ Empty Resource Groups: {e}")
            category_results["Empty Resource Groups"] = []

    # ── Post-query filters: subscription exclusions + DoNotDelete tag ────
    category_results, excl_dropped = apply_subscription_exclusions(
        category_results, excluded)
    if excl_dropped:
        print(f"  {excl_dropped} resource(s) dropped from excluded subscription(s)")
    category_results, dnd_dropped = filter_do_not_delete(category_results)
    if dnd_dropped:
        print(f"  {dnd_dropped} resource(s) skipped (DoNotDelete tag)")

    # ── Cost Management enrichment (real cost data) ──────────────────────
    cost_result = None
    if not args.no_cost_data:
        try:
            from azure_cost_enrichment import enrich_costs

            subs_with_orphans = {
                r.get("subscriptionId", "")
                for rows in category_results.values()
                for r in rows
                if r.get("subscriptionId")
            }
            if subs_with_orphans:
                print("\nEnriching with Cost Management data...")
                cost_result = enrich_costs(
                    credential,
                    sorted(subs_with_orphans),
                    cache_dir=args.cost_cache_dir,
                    use_cache=not args.refresh_cost_data,
                )
                window = cost_result.rolling30d_window
                if window[0]:
                    print(f"  Rolling 30d window: {window[0]} → {window[1]}")
                print(
                    f"  {len(cost_result.subscriptions_queried)} subscription(s) "
                    f"queried, {len(cost_result.cost_map)} resource records, "
                    f"${cost_result.total_rolling30d:,.2f} total rolling 30d spend"
                )
                if cost_result.subscriptions_failed:
                    for sid, err in cost_result.subscriptions_failed.items():
                        name = sub_names.get(sid, sid)
                        print(f"  ⚠  {name}: {err}")
        except ImportError:
            print("  azure_cost_enrichment module not available — using estimates.")
            cost_result = None
        except Exception as e:
            print(f"  Cost enrichment failed ({e}) — falling back to estimates.")
            logger.warning(f"Cost enrichment failed: {e}")
            cost_result = None

    # ── Build ordered list of (category, config, rows) ────────────────────
    all_rows = []
    for cat_name, cfg in QUERIES.items():
        rows = category_results.get(cat_name, [])
        if rows:
            all_rows.append((cat_name, cfg, rows))

    empty_rgs = category_results.get("Empty Resource Groups", [])
    if empty_rgs:
        all_rows.append(("Empty Resource Groups", {"cost": False, "section": "OTHER"}, empty_rgs))

    # ── Counts ────────────────────────────────────────────────────────────
    total = sum(len(r) for _, _, r in all_rows)
    prod_total = sum(
        1 for _, _, rows in all_rows for r in rows
        if classify_resource(r, sub_envs) == "PRODUCTION"
    )
    nonprod_total = total - prod_total
    cost_total = sum(
        1 for _, cfg, rows in all_rows for r in rows
        if cfg["cost"]
    )

    print(f"\nTotal orphaned resources: {total}")
    print(f"  Production: {prod_total}")
    print(f"  Dev/QA/UAT: {nonprod_total}")
    print(f"  Cost-incurring: {cost_total}")

    # ── Generate Excel ────────────────────────────────────────────────────
    print(f"\nGenerating {args.output}...")
    wb = Workbook()

    # ── Summary sheet ─────────────────────────────────────────────────────
    ws_sum = wb.active
    ws_sum.title = "Summary"

    ws_sum.merge_cells("A1:D1")
    title_cell = ws_sum.cell(row=1, column=1, value="Azure Orphaned Resources Report")
    title_cell.font = Font(name="Calibri", bold=True, size=16, color="1F4E79")

    now = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")
    ws_sum.cell(row=2, column=1, value=f"Generated: {now}").font = Font(name="Calibri", size=10, italic=True)
    ws_sum.cell(row=3, column=1, value=f"Scope: {len(sub_names)} subscriptions (tenant-wide)").font = Font(name="Calibri", size=10, italic=True)

    # Cost Management footnote
    if cost_result is not None:
        window_from, window_to = cost_result.rolling30d_window
        ws_sum.cell(
            row=4, column=1,
            value=f"Cost data: Azure Cost Management AmortizedCost — "
                  f"rolling 30d window {window_from} → {window_to}"
        ).font = Font(name="Calibri", size=10, italic=True, color="548235")
    else:
        ws_sum.cell(
            row=4, column=1,
            value="Cost data: hardcoded per-category estimates "
                  "(Cost Management enrichment disabled or unavailable)"
        ).font = Font(name="Calibri", size=10, italic=True, color="C00000")

    # Summary table
    sum_headers = ["Category", "Section", "Production", "Dev/QA/UAT", "Total",
                   "Incurs Cost?", "Monthly Cost", "Rolling 30d (actual)",
                   "Last Billing Month"]
    for col, h in enumerate(sum_headers, 1):
        cell = ws_sum.cell(row=5, column=col, value=h)
        cell.font = SUMMARY_HEADER_FONT
        cell.fill = SUMMARY_HEADER_FILL

    row_num = 6
    grand_prod = 0
    grand_nonprod = 0
    grand_monthly = 0.0
    grand_rolling = 0.0
    grand_last_month = 0.0
    for cat_name, cfg, rows in all_rows:
        prod_count = sum(1 for r in rows if classify_resource(r, sub_envs) == "PRODUCTION")
        nonprod_count = len(rows) - prod_count
        grand_prod += prod_count
        grand_nonprod += nonprod_count

        # Sum per-resource monthly/rolling/last-month rather than
        # category_estimate * count — this picks up real costs where available.
        # Round per row, matching exactly what the detail cells display, so
        # the summary equals the visible column sums to the cent.
        cat_monthly = 0.0
        cat_rolling = 0.0
        cat_last_month = 0.0
        for r in rows:
            monthly, rolling, last_month, _src = _lookup_cost(r, cat_name, cost_result)
            cat_monthly += round(monthly, 2)
            cat_rolling += round(rolling, 2)
            cat_last_month += round(last_month, 2)
        grand_monthly += cat_monthly
        grand_rolling += cat_rolling
        grand_last_month += cat_last_month

        ws_sum.cell(row=row_num, column=1, value=cat_name)
        ws_sum.cell(row=row_num, column=2, value=cfg["section"])
        ws_sum.cell(row=row_num, column=3, value=prod_count)
        ws_sum.cell(row=row_num, column=4, value=nonprod_count)
        ws_sum.cell(row=row_num, column=5, value=len(rows)).font = Font(bold=True)
        cost_cell = ws_sum.cell(row=row_num, column=6, value="Yes" if cfg["cost"] else "No")
        cost_cell.font = COST_YES if cfg["cost"] else COST_NO

        monthly_cell = ws_sum.cell(row=row_num, column=7, value=round(cat_monthly, 2))
        monthly_cell.number_format = '$#,##0.00'
        rolling_cell = ws_sum.cell(row=row_num, column=8, value=round(cat_rolling, 2))
        rolling_cell.number_format = '$#,##0.00'
        last_cell = ws_sum.cell(row=row_num, column=9, value=round(cat_last_month, 2))
        last_cell.number_format = '$#,##0.00'

        for c in range(1, 10):
            ws_sum.cell(row=row_num, column=c).border = THIN_BORDER
        row_num += 1

    # Totals row
    row_num += 1
    ws_sum.cell(row=row_num, column=1, value="TOTAL").font = Font(bold=True, size=12)
    ws_sum.cell(row=row_num, column=3, value=grand_prod).font = Font(bold=True, size=12)
    ws_sum.cell(row=row_num, column=4, value=grand_nonprod).font = Font(bold=True, size=12)
    ws_sum.cell(row=row_num, column=5, value=total).font = Font(bold=True, size=12)
    grand_monthly_cell = ws_sum.cell(row=row_num, column=7, value=round(grand_monthly, 2))
    grand_monthly_cell.font = Font(bold=True, size=12)
    grand_monthly_cell.number_format = '$#,##0.00'
    grand_rolling_cell = ws_sum.cell(row=row_num, column=8, value=round(grand_rolling, 2))
    grand_rolling_cell.font = Font(bold=True, size=12)
    grand_rolling_cell.number_format = '$#,##0.00'
    grand_last_cell = ws_sum.cell(row=row_num, column=9, value=round(grand_last_month, 2))
    grand_last_cell.font = Font(bold=True, size=12)
    grand_last_cell.number_format = '$#,##0.00'

    auto_width(ws_sum)

    # ── Production detail sheet ───────────────────────────────────────────
    ws_prod = wb.create_sheet("Production & SharedServices")
    write_detail_sheet(ws_prod, "Production", all_rows, sub_names, sub_envs, "PRODUCTION", cost_result)

    # ── Dev/QA/UAT detail sheet ───────────────────────────────────────────
    ws_dev = wb.create_sheet("Dev - QA - UAT")
    write_detail_sheet(ws_dev, "Dev/QA/UAT", all_rows, sub_names, sub_envs, "NON-PRODUCTION", cost_result)

    # ── All Resources sheet (flat, filterable) ────────────────────────────
    ws_all = wb.create_sheet("All Resources")
    headers = ["Category", "Section", "Resource Name", "Resource Group",
               "Location", "Subscription", "SKU / Detail", "Incurs Cost?",
               "Monthly Cost", "Rolling 30d (actual)", "Last Billing Month",
               "Cost Source", "Environment", "Tags"]
    for col, h in enumerate(headers, 1):
        cell = ws_all.cell(row=1, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
    ws_all.freeze_panes = "A2"
    ws_all.auto_filter.ref = "A1:N1"

    row_num = 2
    for cat_name, cfg, resources in all_rows:
        for r in resources:
            sub_id = r.get("subscriptionId", "")
            env = classify_resource(r, sub_envs)
            monthly, rolling, last_month, source = _lookup_cost(r, cat_name, cost_result)

            ws_all.cell(row=row_num, column=1, value=cat_name)
            ws_all.cell(row=row_num, column=2, value=cfg["section"])
            ws_all.cell(row=row_num, column=3, value=r.get("name", ""))
            ws_all.cell(row=row_num, column=4, value=r.get("resourceGroup", ""))
            ws_all.cell(row=row_num, column=5, value=r.get("location", ""))
            ws_all.cell(row=row_num, column=6, value=sub_names.get(sub_id, sub_id))
            ws_all.cell(row=row_num, column=7, value=r.get("sku", "") or "")

            cost_cell = ws_all.cell(row=row_num, column=8, value="Yes" if cfg["cost"] else "No")
            cost_cell.font = COST_YES if cfg["cost"] else COST_NO

            monthly_cell = ws_all.cell(row=row_num, column=9, value=round(monthly, 2))
            monthly_cell.number_format = '$#,##0.00'
            rolling_cell = ws_all.cell(row=row_num, column=10, value=round(rolling, 2))
            rolling_cell.number_format = '$#,##0.00'
            last_cell = ws_all.cell(row=row_num, column=11, value=round(last_month, 2))
            last_cell.number_format = '$#,##0.00'
            ws_all.cell(row=row_num, column=12, value=source)

            env_cell = ws_all.cell(row=row_num, column=13, value=env)
            env_cell.fill = PROD_FILL if env == "PRODUCTION" else NONPROD_FILL
            ws_all.cell(row=row_num, column=14, value=_format_tags(r.get("tags")))

            for c in range(1, 15):
                ws_all.cell(row=row_num, column=c).border = THIN_BORDER
            row_num += 1

    auto_width(ws_all)

    # ── Save ──────────────────────────────────────────────────────────────
    wb.save(args.output)
    print(f"\n✓ Report saved to: {args.output}")
    print(f"  Sheets: Summary | Production & SharedServices | Dev - QA - UAT | All Resources")
    if cost_result is not None:
        total_rows = sum(len(rows) for _, _, rows in all_rows)
        real_count = 0
        for cat_name, _, rows in all_rows:
            for r in rows:
                _, _, _, src = _lookup_cost(r, cat_name, cost_result)
                if src == "costManagement":
                    real_count += 1
        print(f"  {real_count} of {total_rows} rows backed by real Cost Management data")
    return 0


if __name__ == "__main__":
    sys.exit(main() or 0)
